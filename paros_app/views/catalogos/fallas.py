import csv
import io
import openpyxl

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_http_methods

from ...models import Area, CatalogoFalla
from ...views.utils import _excel_response, _estilo_cabecera
from login_app.permisos import permiso_requerido, get_perfil


@login_required
@permiso_requerido('ver_catalogos')
def catalogo_fallas_general(request):
    perfil = get_perfil(request.user)
    areas  = Area.objects.prefetch_related('catalogo_fallas').all()
    if not (request.user.is_superuser or (perfil and perfil.es_admin)):
        if perfil and perfil.areas_permitidas.exists():
            areas = areas.filter(id__in=perfil.areas_permitidas.all())
        else:
            areas = areas.none()
    puede_gestionar = request.user.is_superuser or (perfil and (perfil.es_admin or perfil.gestionar_catalogos))
    return render(request, 'paros_app/catalogo_fallas_general.html', {
        'areas':           areas,
        'puede_gestionar': puede_gestionar,
    })


@login_required
@permiso_requerido('ver_catalogos')
def catalogo_fallas(request, area_id):
    perfil = get_perfil(request.user)
    if not (request.user.is_superuser or (perfil and perfil.es_admin)):
        if perfil and perfil.areas_permitidas.exists():
            if not perfil.areas_permitidas.filter(id=area_id).exists():
                messages.error(request, "No tienes acceso al catálogo de esta área.")
                return redirect('catalogo_fallas_general')
    area = get_object_or_404(Area, id=area_id)
    qs   = CatalogoFalla.objects.filter(area=area).order_by('codigo')
    q    = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(Q(nombre__icontains=q) | Q(codigo__icontains=q))
    paginator = Paginator(qs, 50)
    fallas    = paginator.get_page(request.GET.get('page', 1))
    puede_gestionar = request.user.is_superuser or (perfil and (perfil.es_admin or perfil.gestionar_catalogos))
    return render(request, 'paros_app/catalogo_fallas.html', {
        'area':            area,
        'fallas':          fallas,
        'page_obj':        fallas,
        'areas':           Area.objects.all(),
        'q':               q,
        'puede_gestionar': puede_gestionar,
    })


@login_required
@require_http_methods(["POST"])
def eliminar_falla(request, falla_id):
    falla   = get_object_or_404(CatalogoFalla, id=falla_id)
    area_id = falla.area_id
    falla.delete()
    return redirect('catalogo_fallas', area_id=area_id)


@login_required
@permiso_requerido('gestionar_catalogos')
@require_http_methods(["POST"])
def limpiar_fallas_area(request, area_id):
    area = get_object_or_404(Area, id=area_id)
    CatalogoFalla.objects.filter(area=area).delete()
    messages.success(request, f"Catálogo de fallas de '{area.nombre}' eliminado.")
    return redirect('catalogo_fallas_general')


@login_required
@permiso_requerido('gestionar_catalogos')
def importar_fallas_v2(request):
    """Importación con 3 columnas en Excel: código | área | falla."""
    errores = []
    resumen = None

    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        modo    = request.POST.get('modo', 'agregar')
        if not archivo:
            errores.append("Selecciona un archivo Excel.")
        else:
            nombre = archivo.name.lower()
            filas  = []
            try:
                if nombre.endswith(('.xlsx', '.xls')):
                    wb   = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
                    ws   = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    for i, row in enumerate(rows[1:], start=2):
                        if not any(row): continue
                        codigo      = str(row[0] or '').strip() if len(row) > 0 else ''
                        nombre_area = str(row[1] or '').strip() if len(row) > 1 else ''
                        falla_nom   = str(row[2] or '').strip() if len(row) > 2 else ''
                        area_origen = str(row[3] or '').strip() if len(row) > 3 else ''
                        if not codigo or not nombre_area or not falla_nom:
                            errores.append(f"Fila {i}: código, área y falla son obligatorios.")
                            continue
                        filas.append((codigo, nombre_area, falla_nom, area_origen))
                elif nombre.endswith('.csv'):
                    contenido = archivo.read().decode('utf-8-sig')
                    reader    = csv.DictReader(io.StringIO(contenido))
                    for i, row in enumerate(reader, start=2):
                        codigo      = (row.get('codigo') or row.get('código') or '').strip()
                        nombre_area = (row.get('area')   or row.get('área')   or '').strip()
                        falla_nom   = (row.get('falla')  or '').strip()
                        area_origen = (row.get('area_origen') or row.get('area origen') or '').strip()
                        if not codigo or not nombre_area or not falla_nom:
                            errores.append(f"Fila {i}: datos incompletos.")
                            continue
                        filas.append((codigo, nombre_area, falla_nom, area_origen))
                else:
                    errores.append("Solo se aceptan archivos .xlsx, .xls o .csv")
            except Exception as e:
                errores.append(f"Error al leer el archivo: {e}")

            if filas and not errores:
                creados = actualizados = omitidos = 0
                areas_borradas   = set()
                errores_guardado = []
                try:
                    with transaction.atomic():
                        for codigo, nombre_area, falla_nom, area_origen in filas:
                            try:
                                area = Area.objects.get(nombre__iexact=nombre_area)
                            except Area.DoesNotExist:
                                errores_guardado.append(f"Área '{nombre_area}' no existe.")
                                omitidos += 1
                                continue
                            if modo == 'reemplazar' and area.id not in areas_borradas:
                                CatalogoFalla.objects.filter(area=area).delete()
                                areas_borradas.add(area.id)
                            _, created = CatalogoFalla.objects.update_or_create(
                                area=area, codigo=codigo,
                                defaults={'nombre': falla_nom, 'descripcion': '', 'area_origen': area_origen}
                            )
                            if created: creados += 1
                            else: actualizados += 1
                        if errores_guardado:
                            raise Exception("Errores en importación")
                except Exception:
                    errores.extend(errores_guardado)
                    creados = actualizados = 0

                if not errores:
                    resumen = {
                        'creados':      creados,
                        'actualizados': actualizados,
                        'omitidos':     omitidos,
                        'total':        creados + actualizados,
                    }

    areas = Area.objects.all()
    return render(request, 'paros_app/importar_fallas_v2.html', {
        'areas': areas, 'errores': errores, 'resumen': resumen,
    })


@login_required
@permiso_requerido('gestionar_catalogos')
def importar_fallas_por_area(request, area_id):
    area    = get_object_or_404(Area, id=area_id)
    errores = []
    resumen = None

    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        modo    = request.POST.get('modo', 'agregar')
        if not archivo:
            errores.append("Selecciona un archivo.")
        else:
            filas = []
            try:
                nombre = archivo.name.lower()
                if nombre.endswith(('.xlsx', '.xls')):
                    wb   = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
                    ws   = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        errores.append("El archivo está vacío.")
                    else:
                        headers = [str(c).strip().lower() if c else '' for c in rows[0]]
                        def col(names):
                            for n in names:
                                if n in headers: return headers.index(n)
                            return None
                        idx_cod  = col(['código', 'codigo'])
                        idx_fall = col(['falla'])
                        idx_orig = col(['área de origen', 'area de origen', 'area_origen'])
                        idx_area = col(['área', 'area'])
                        if idx_cod is None or idx_fall is None:
                            errores.append("El archivo debe tener columnas 'Código' y 'Falla'.")
                        else:
                            for i, row in enumerate(rows[1:], start=2):
                                if not any(row): continue
                                if idx_area is not None and idx_area < len(row):
                                    area_col = str(row[idx_area] or '').strip()
                                    if area_col and area_col.lower() != area.nombre.lower():
                                        continue
                                codigo      = str(row[idx_cod]  or '').strip()
                                falla_nom   = str(row[idx_fall] or '').strip()
                                area_origen = str(row[idx_orig] or '').strip() if idx_orig is not None and idx_orig < len(row) else ''
                                if not codigo or not falla_nom:
                                    errores.append(f"Fila {i}: código y falla son obligatorios.")
                                    continue
                                filas.append((codigo, falla_nom, area_origen))
                elif nombre.endswith('.csv'):
                    contenido = archivo.read().decode('utf-8-sig')
                    reader = csv.DictReader(io.StringIO(contenido))
                    for i, row in enumerate(reader, start=2):
                        row_lower = {k.strip().lower(): v for k, v in row.items()}
                        area_col = (row_lower.get('área') or row_lower.get('area') or '').strip()
                        if area_col and area_col.lower() != area.nombre.lower():
                            continue
                        codigo      = (row_lower.get('código') or row_lower.get('codigo') or '').strip()
                        falla_nom   = (row_lower.get('falla') or '').strip()
                        area_origen = (row_lower.get('área de origen') or row_lower.get('area de origen') or row_lower.get('area_origen') or '').strip()
                        if not codigo or not falla_nom:
                            errores.append(f"Fila {i}: datos incompletos.")
                            continue
                        filas.append((codigo, falla_nom, area_origen))
                else:
                    errores.append("Solo .xlsx, .xls o .csv")
            except Exception as e:
                errores.append(f"Error al leer el archivo: {e}")

            if filas and not errores:
                if modo == 'reemplazar':
                    CatalogoFalla.objects.filter(area=area).delete()
                creados = actualizados = 0
                for codigo, falla_nom, area_origen in filas:
                    _, created = CatalogoFalla.objects.update_or_create(
                        area=area, codigo=codigo,
                        defaults={'nombre': falla_nom, 'descripcion': '', 'area_origen': area_origen}
                    )
                    if created: creados += 1
                    else: actualizados += 1
                resumen = {'creados': creados, 'actualizados': actualizados, 'area': area.nombre}

    return render(request, 'paros_app/importar_fallas_por_area.html', {
        'area': area, 'errores': errores, 'resumen': resumen,
    })


@login_required
def descargar_plantilla_fallas_v2(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="plantilla_fallas.csv"'
    response.write('\ufeff')
    writer = csv.writer(response)
    writer.writerow(['codigo', 'area', 'falla'])
    writer.writerow(['F-001', 'Producción', 'Falla mecánica'])
    writer.writerow(['F-002', 'Producción', 'Falla eléctrica'])
    writer.writerow(['F-003', 'Mantenimiento', 'Desgaste de componentes'])
    return response


@login_required
@permiso_requerido('ver_catalogos')
def exportar_fallas(request, area_id=None):
    qs = CatalogoFalla.objects.select_related('area').all()
    if area_id:
        area  = get_object_or_404(Area, id=area_id)
        qs    = qs.filter(area=area)
        fname = f'fallas_{area.nombre.replace(" ", "_")}.xlsx'
    else:
        fname = 'fallas_general.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Fallas'
    _estilo_cabecera(ws, ['Código', 'Área', 'Falla', 'Área de origen'], [14, 22, 38, 20])
    for f in qs.order_by('area__nombre', 'codigo'):
        ws.append([f.codigo, f.area.nombre, f.nombre, f.area_origen or ''])
    response = _excel_response(fname)
    wb.save(response)
    return response


@login_required
@permiso_requerido('gestionar_catalogos')
def agregar_falla(request, area_id):
    area = get_object_or_404(Area, id=area_id)
    if request.method == 'POST':
        codigo      = request.POST.get('codigo', '').strip()
        nombre      = request.POST.get('nombre', '').strip()
        area_origen = request.POST.get('area_origen', '').strip()

        if not codigo or not nombre:
            messages.error(request, "Código y nombre son obligatorios.")
        elif CatalogoFalla.objects.filter(area=area, codigo=codigo).exists():
            messages.error(request, f"Ya existe una falla con el código '{codigo}' en esta área.")
        elif CatalogoFalla.objects.filter(area=area, nombre__iexact=nombre).exists():
            messages.error(request, f"Ya existe una falla con el nombre '{nombre}' en esta área.")
        else:
            CatalogoFalla.objects.create(
                area=area,
                codigo=codigo,
                nombre=nombre,
                descripcion='',
                area_origen=area_origen,
            )
            messages.success(request, f"Falla '{nombre}' agregada correctamente.")

    return redirect('catalogo_fallas_general')