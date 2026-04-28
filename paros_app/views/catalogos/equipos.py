import csv
import io
import openpyxl

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_http_methods

from ...models import Area, CatalogoEquipo
from ...views.utils import _excel_response, _estilo_cabecera
from login_app.permisos import permiso_requerido, get_perfil


@login_required
@permiso_requerido('ver_catalogos')
def catalogo_equipos_general(request):
    perfil = get_perfil(request.user)
    q     = request.GET.get('q', '').strip()
    areas = Area.objects.prefetch_related('catalogo_equipos').all()
    if not (request.user.is_superuser or (perfil and perfil.es_admin)):
        if perfil and perfil.areas_permitidas.exists():
            areas = areas.filter(id__in=perfil.areas_permitidas.all())
        else:
            areas = areas.none()
    page_obj = None
    if q:
        qs = CatalogoEquipo.objects.select_related('area').order_by('area__nombre', 'codigo')
        if not (request.user.is_superuser or (perfil and perfil.es_admin)):
            if perfil and perfil.areas_permitidas.exists():
                qs = qs.filter(area__in=perfil.areas_permitidas.all())
        qs = qs.filter(Q(equipo__icontains=q) | Q(codigo__icontains=q))
        paginator = Paginator(qs, 50)
        page_obj  = paginator.get_page(request.GET.get('page', 1))
    puede_gestionar = request.user.is_superuser or (perfil and (perfil.es_admin or perfil.gestionar_catalogos))
    return render(request, 'paros_app/catalogo_equipos.html', {
        'page_obj': page_obj,
        'areas':    areas,
        'q':        q,
        'puede_gestionar': puede_gestionar,
    })


@login_required
@permiso_requerido('gestionar_catalogos')
@require_http_methods(["POST"])
def limpiar_equipos_area(request, area_id):
    area = get_object_or_404(Area, id=area_id)
    CatalogoEquipo.objects.filter(area=area).delete()
    messages.success(request, f"Catálogo de equipos de '{area.nombre}' eliminado.")
    return redirect('catalogo_equipos')


@login_required
@permiso_requerido('gestionar_catalogos')
def importar_equipos(request):
    errores = []
    resumen = None

    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        modo    = request.POST.get('modo', 'agregar')
        if not archivo:
            errores.append("Selecciona un archivo Excel.")
        else:
            nombre = archivo.name.lower()
            filas  = []
            try:
                if nombre.endswith(('.xlsx', '.xls')):
                    wb   = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
                    ws   = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    for i, row in enumerate(rows[1:], start=2):
                        if not any(row): continue
                        codigo      = str(row[0] or '').strip() if len(row) > 0 else ''
                        nombre_area = str(row[1] or '').strip() if len(row) > 1 else ''
                        equipo_nom  = str(row[2] or '').strip() if len(row) > 2 else ''
                        if not codigo or not nombre_area or not equipo_nom:
                            errores.append(f"Fila {i}: código, área y equipo son obligatorios.")
                            continue
                        filas.append((codigo, nombre_area, equipo_nom))
                elif nombre.endswith('.csv'):
                    contenido = archivo.read().decode('utf-8-sig')
                    reader    = csv.DictReader(io.StringIO(contenido))
                    for i, row in enumerate(reader, start=2):
                        codigo      = (row.get('codigo') or row.get('código') or '').strip()
                        nombre_area = (row.get('area')   or row.get('área')   or '').strip()
                        equipo_nom  = (row.get('equipo') or '').strip()
                        if not codigo or not nombre_area or not equipo_nom:
                            errores.append(f"Fila {i}: datos incompletos.")
                            continue
                        filas.append((codigo, nombre_area, equipo_nom))
                else:
                    errores.append("Solo se aceptan .xlsx, .xls o .csv")
            except Exception as e:
                errores.append(f"Error al leer el archivo: {e}")

            if filas and not errores:
                creados = actualizados = omitidos = 0
                areas_borradas   = set()
                errores_guardado = []
                try:
                    with transaction.atomic():
                        for codigo, nombre_area, equipo_nom in filas:
                            try:
                                area = Area.objects.get(nombre__iexact=nombre_area)
                            except Area.DoesNotExist:
                                errores_guardado.append(f"Área '{nombre_area}' no existe.")
                                omitidos += 1
                                continue
                            if modo == 'reemplazar' and area.id not in areas_borradas:
                                CatalogoEquipo.objects.filter(area=area).delete()
                                areas_borradas.add(area.id)
                            _, created = CatalogoEquipo.objects.update_or_create(
                                area=area, codigo=codigo,
                                defaults={'equipo': equipo_nom}
                            )
                            if created: creados += 1
                            else: actualizados += 1
                        if errores_guardado:
                            raise Exception("Errores en importación")
                except Exception:
                    errores.extend(errores_guardado)
                    creados = actualizados = 0

                if not errores:
                    resumen = {
                        'creados':      creados,
                        'actualizados': actualizados,
                        'omitidos':     omitidos,
                        'total':        creados + actualizados,
                    }

    areas = Area.objects.all()
    return render(request, 'paros_app/importar_equipos.html', {
        'areas': areas, 'errores': errores, 'resumen': resumen,
    })


@login_required
@permiso_requerido('gestionar_catalogos')
def importar_equipos_por_area(request, area_id):
    area    = get_object_or_404(Area, id=area_id)
    errores = []
    resumen = None

    if request.method == 'POST':
        archivo = request.FILES.get('archivo')
        modo    = request.POST.get('modo', 'agregar')
        if not archivo:
            errores.append("Selecciona un archivo.")
        else:
            filas = []
            try:
                nombre = archivo.name.lower()
                if nombre.endswith(('.xlsx', '.xls')):
                    wb   = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
                    ws   = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        errores.append("El archivo está vacío.")
                    else:
                        headers = [str(c).strip().lower() if c else '' for c in rows[0]]
                        def col_eq(names):
                            for n in names:
                                if n in headers: return headers.index(n)
                            return None
                        idx_cod  = col_eq(['código', 'codigo'])
                        idx_eq   = col_eq(['equipo'])
                        idx_area = col_eq(['área', 'area'])
                        if idx_cod is None or idx_eq is None:
                            errores.append("El archivo debe tener columnas 'Código' y 'Equipo'.")
                        else:
                            for i, row in enumerate(rows[1:], start=2):
                                if not any(row): continue
                                if idx_area is not None and idx_area < len(row):
                                    area_col = str(row[idx_area] or '').strip()
                                    if area_col and area_col.lower() != area.nombre.lower():
                                        continue
                                codigo     = str(row[idx_cod] or '').strip()
                                equipo_nom = str(row[idx_eq]  or '').strip()
                                if not codigo or not equipo_nom:
                                    errores.append(f"Fila {i}: código y equipo son obligatorios.")
                                    continue
                                filas.append((codigo, equipo_nom))
                elif nombre.endswith('.csv'):
                    contenido = archivo.read().decode('utf-8-sig')
                    reader = csv.DictReader(io.StringIO(contenido))
                    for i, row in enumerate(reader, start=2):
                        row_lower = {k.strip().lower(): v for k, v in row.items()}
                        area_col = (row_lower.get('área') or row_lower.get('area') or '').strip()
                        if area_col and area_col.lower() != area.nombre.lower():
                            continue
                        codigo     = (row_lower.get('código') or row_lower.get('codigo') or '').strip()
                        equipo_nom = (row_lower.get('equipo') or '').strip()
                        if not codigo or not equipo_nom:
                            errores.append(f"Fila {i}: datos incompletos.")
                            continue
                        filas.append((codigo, equipo_nom))
                else:
                    errores.append("Solo .xlsx, .xls o .csv")
            except Exception as e:
                errores.append(f"Error al leer el archivo: {e}")

            if filas and not errores:
                if modo == 'reemplazar':
                    CatalogoEquipo.objects.filter(area=area).delete()
                creados = actualizados = 0
                for codigo, equipo_nom in filas:
                    _, created = CatalogoEquipo.objects.update_or_create(
                        area=area, codigo=codigo,
                        defaults={'equipo': equipo_nom}
                    )
                    if created: creados += 1
                    else: actualizados += 1
                resumen = {'creados': creados, 'actualizados': actualizados, 'area': area.nombre}

    return render(request, 'paros_app/importar_equipos_por_area.html', {
        'area': area, 'errores': errores, 'resumen': resumen,
    })


@login_required
def descargar_plantilla_equipos(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="plantilla_equipos.csv"'
    response.write('\ufeff')
    writer = csv.writer(response)
    writer.writerow(['codigo', 'area', 'equipo'])
    writer.writerow(['EQ-001', 'Producción', 'Compresor A'])
    writer.writerow(['EQ-002', 'Producción', 'Bomba hidráulica 1'])
    return response


@login_required
@require_http_methods(["POST"])
def eliminar_equipo(request, equipo_id):
    equipo = get_object_or_404(CatalogoEquipo, id=equipo_id)
    equipo.delete()
    return redirect('catalogo_equipos')


@login_required
@permiso_requerido('ver_catalogos')
def exportar_equipos(request, area_id=None):
    qs = CatalogoEquipo.objects.select_related('area').all()
    if area_id:
        area  = get_object_or_404(Area, id=area_id)
        qs    = qs.filter(area=area)
        fname = f'equipos_{area.nombre.replace(" ", "_")}.xlsx'
    else:
        fname = 'equipos_general.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Equipos'
    _estilo_cabecera(ws, ['Código', 'Área', 'Equipo'], [14, 22, 38])
    for eq in qs.order_by('area__nombre', 'codigo'):
        ws.append([eq.codigo, eq.area.nombre, eq.equipo])
    response = _excel_response(fname)
    wb.save(response)
    return response

@login_required
@permiso_requerido('gestionar_catalogos')
def agregar_equipos(request, area_id):
    area = get_object_or_404(Area, id=area_id)
    if request.method == 'POST':
        codigo     = request.POST.get('codigo', '').strip()
        equipo_nom = request.POST.get('equipo', '').strip()
        if not codigo or not equipo_nom:
            messages.error(request, "Código y equipo son obligatorios.")
        elif CatalogoEquipo.objects.filter(area=area, codigo=codigo).exists():
            messages.error(request, f"Ya existe un equipo con el código '{codigo}' en esta área.")
        elif CatalogoEquipo.objects.filter(area=area, equipo__iexact=equipo_nom).exists():
            messages.error(request, f"Ya existe un equipo con el nombre '{equipo_nom}' en esta área.")
        else:
            CatalogoEquipo.objects.create(area=area, codigo=codigo, equipo=equipo_nom)
            messages.success(request, f"Equipo '{equipo_nom}' agregado correctamente.")
    return redirect('catalogo_equipos')