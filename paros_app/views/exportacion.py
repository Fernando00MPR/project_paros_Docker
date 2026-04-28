import csv
import json
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import redirect, render

from ..models import Area, Paro
from login_app.permisos import get_perfil
from .utils import _aplicar_filtros, _parse_fecha


@login_required
def exportar_csv(request):
    from datetime import date as _date
    perfil = get_perfil(request.user)
    if not (request.user.is_superuser or (perfil and (perfil.es_admin or perfil.exportar_paros))):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('Sin permiso para exportar.')

    get_params = request.GET.copy()
    if not get_params.get('fecha_desde'):
        get_params['fecha_desde'] = _date.today().strftime('%Y-%m-%d')
    if not get_params.get('fecha_hasta'):
        get_params['fecha_hasta'] = _date.today().strftime('%Y-%m-%d')

    paros_qs = _aplicar_filtros(Paro.objects.select_related('area').all(), get_params)
    area_id = request.GET.get('area_id')
    if area_id:
        paros_qs = paros_qs.filter(area_id=area_id)
        nombre_area = Area.objects.filter(id=area_id).first()
        filename = f"Paros_{nombre_area.nombre.replace(' ','_')}.csv" if nombre_area else "Paros.csv"
    else:
        filename = "Paros_Todas_Las_Areas.csv"

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('\ufeff')
    writer = csv.writer(response)
    ESTATUS_L = {'rojo': 'Sin revisar', 'amarillo': 'Pendiente', 'verde': 'Revisado'}
    writer.writerow(['Área', 'Fecha', 'Sem.', 'Turno', 'Falla', 'Responsable', 'Equipo', 'Hora', 'Tiempo (min)', 'Estatus', 'Comentarios'])
    for p in paros_qs:
        writer.writerow([
            p.area.nombre, p.fecha.strftime('%d/%m/%Y'),
            p.fecha.isocalendar()[1],
            p.get_turno_display(), p.falla, p.responsable, p.equipo,
            p.hora.strftime('%H:%M'), p.tiempo_minutos,
            ESTATUS_L.get(p.estatus, p.estatus), p.comentarios,
        ])
    return response


@login_required
def exportar_excel(request):
    from datetime import date as _date
    perfil = get_perfil(request.user)
    if not (request.user.is_superuser or (perfil and (perfil.es_admin or perfil.exportar_paros))):
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden('Sin permiso para exportar.')

    get_params = request.GET.copy()
    if not get_params.get('fecha_desde'):
        get_params['fecha_desde'] = _date.today().strftime('%Y-%m-%d')
    if not get_params.get('fecha_hasta'):
        get_params['fecha_hasta'] = _date.today().strftime('%Y-%m-%d')

    paros_qs = _aplicar_filtros(Paro.objects.select_related('area').all(), get_params)
    area_id = request.GET.get('area_id')
    if area_id:
        paros_qs = paros_qs.filter(area_id=area_id)
        nombre_area = Area.objects.filter(id=area_id).first()
        filename = f"Paros_{nombre_area.nombre.replace(' ','_')}.xlsx" if nombre_area else "Paros.xlsx"
    else:
        filename = "Paros_Todas_Las_Areas.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Paros'
    ESTATUS_L = {'rojo': 'Sin revisar', 'amarillo': 'Pendiente', 'verde': 'Revisado'}
    cabeceras = ['Área', 'Fecha', 'Sem.', 'Turno', 'Falla', 'Responsable', 'Equipo', 'Hora', 'Tiempo (min)', 'Estatus', 'Comentarios']
    ws.append(cabeceras)

    header_fill = PatternFill('solid', fgColor='4F46E5')
    header_font = Font(bold=True, color='FFFFFF')
    center = Alignment(horizontal='center', vertical='center')
    for col in range(1, len(cabeceras) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for p in paros_qs:
        ws.append([
            p.area.nombre, p.fecha.strftime('%d/%m/%Y'),
            p.fecha.isocalendar()[1],
            p.get_turno_display(), p.falla, p.responsable, p.equipo,
            p.hora.strftime('%H:%M'), p.tiempo_minutos,
            ESTATUS_L.get(p.estatus, p.estatus), p.comentarios or '',
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = center
            if cell.column == 2:
                cell.number_format = '@'

    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ''
                if len(val) > max_len:
                    max_len = len(val)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def importar_paros(request):
    from datetime import datetime as dt
    perfil = get_perfil(request.user)
    es_admin = request.user.is_superuser or (perfil and perfil.es_admin)

    if not (es_admin or (perfil and perfil.importar_paros)):
        messages.error(request, "No tienes permiso para importar paros.")
        return redirect('lista_paros')

    COLUMNAS = ['Área', 'Fecha', 'Sem.', 'Turno', 'Falla', 'Responsable', 'Equipo', 'Hora', 'Tiempo (min)', 'Estatus', 'Comentarios']
    ESTATUS_MAP = {'sin revisar': 'rojo', 'pendiente': 'amarillo', 'revisado': 'verde'}
    TURNO_MAP = {'turno 1': 1, 'turno 2': 2, '1': 1, '2': 2}

    errores = []
    resumen = None
    filas_permitidas = []
    filas_rechazadas = []
    pendiente_confirmacion = False
    datos_sesion = None

    area_id_param = request.GET.get('area_id') or request.POST.get('area_id_param')
    try:
        area_id_param = int(area_id_param) if area_id_param else None
    except (ValueError, TypeError):
        area_id_param = None

    if es_admin:
        if area_id_param:
            areas_permitidas_ids = {area_id_param}
        else:
            areas_permitidas_ids = None
    else:
        user_areas = set(perfil.areas_permitidas.values_list('id', flat=True))
        if area_id_param:
            if area_id_param in user_areas:
                areas_permitidas_ids = {area_id_param}
            else:
                areas_permitidas_ids = set()
        else:
            areas_permitidas_ids = user_areas if user_areas else set()

    def _crear_paros(filas_list):
        importados = 0
        for f in filas_list:
            try:
                area_obj = Area.objects.get(nombre__iexact=f['area'])
                if areas_permitidas_ids is not None and area_obj.id not in areas_permitidas_ids:
                    continue
                Paro.objects.create(
                    area=area_obj,
                    fecha=_parse_fecha(f['fecha']),
                    turno=f['turno'],
                    falla=f['falla'],
                    responsable=f['responsable'],
                    equipo=f['equipo'],
                    hora=dt.strptime(f['hora'], '%H:%M').time(),
                    tiempo_minutos=int(f['tiempo']),
                    estatus=f['estatus'],
                    comentarios=f.get('comentarios', ''),
                )
                importados += 1
            except Exception:
                pass
        return importados

    if request.method == 'POST':
        accion = request.POST.get('accion', '')

        if accion == 'importar_permitidas':
            filas_json = request.POST.get('filas_permitidas', '[]')
            filas = json.loads(filas_json)
            importados = _crear_paros(filas)
            messages.success(request, f"{importados} paro(s) importados correctamente.")
            if area_id_param:
                return redirect('lista_paros_por_area', area_id=area_id_param)
            return redirect('lista_paros')

        archivo = request.FILES.get('archivo')
        if not archivo:
            errores.append("Selecciona un archivo Excel o CSV.")
        else:
            nombre = archivo.name.lower()
            filas_raw = []
            try:
                if nombre.endswith(('.xlsx', '.xls')):
                    wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
                    ws = wb.active
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        errores.append("El archivo está vacío.")
                    else:
                        headers = [str(c).strip() if c else '' for c in rows[0]]
                        faltantes = [c for c in COLUMNAS if c not in headers]
                        if faltantes:
                            errores.append(f"Columnas faltantes: {', '.join(faltantes)}")
                        else:
                            for i, row in enumerate(rows[1:], 2):
                                if not any(row):
                                    continue
                                def _cell_val(val):
                                    from datetime import datetime as _dt, date as _date
                                    if isinstance(val, (_dt, _date)):
                                        if isinstance(val, _dt):
                                            return val.date().strftime('%d/%m/%Y')
                                        return val.strftime('%d/%m/%Y')
                                    return str(val or '').strip()
                                d = {headers[j]: _cell_val(row[j]) for j in range(len(headers))}
                                filas_raw.append((i, d))

                elif nombre.endswith('.csv'):
                    import unicodedata as _ud
                    try:
                        contenido = archivo.read().decode('utf-8-sig')
                    except UnicodeDecodeError:
                        archivo.seek(0)
                        contenido = archivo.read().decode('latin-1')
                    reader = csv.DictReader(io.StringIO(contenido))
                    raw_headers = reader.fieldnames or []

                    def _norm_h(s):
                        s = ''.join(c for c in _ud.normalize('NFD', str(s)) if _ud.category(c) != 'Mn')
                        return s.strip().lower()

                    col_norm = {_norm_h(c): c for c in COLUMNAS}
                    header_map = {rh: col_norm[_norm_h(rh)] for rh in raw_headers if _norm_h(rh) in col_norm}
                    faltantes = [c for c in COLUMNAS if c not in header_map.values()]
                    if faltantes:
                        errores.append(f"Columnas faltantes: {', '.join(faltantes)}")
                    else:
                        for i, row in enumerate(reader, 2):
                            row_norm = {header_map.get(k, k): v for k, v in row.items()}
                            filas_raw.append((i, row_norm))
                else:
                    errores.append("Formato no soportado. Usa Excel (.xlsx) o CSV.")
            except Exception as ex:
                errores.append(f"Error al leer el archivo: {ex}")

            if not errores and filas_raw:
                for i, d in filas_raw:
                    area_nombre = (d.get('Área') or d.get('Area') or d.get('ÁREA') or d.get('AREA') or '').strip()
                    try:
                        area = Area.objects.get(nombre__iexact=area_nombre)
                    except Area.DoesNotExist:
                        import unicodedata
                        def sin_acentos(s):
                            return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
                        area = None
                        for a in Area.objects.all():
                            if sin_acentos(a.nombre).lower() == sin_acentos(area_nombre).lower():
                                area = a
                                break
                        if area is None:
                            errores.append(f"Fila {i}: área '{area_nombre}' no existe en el sistema.")
                            continue

                    permitida = es_admin or (areas_permitidas_ids is None) or (area.id in areas_permitidas_ids)
                    fila_data = {
                        'area':        area_nombre,
                        'fecha':       d.get('Fecha', ''),
                        'turno':       TURNO_MAP.get(d.get('Turno', '').lower(), 1),
                        'falla':       d.get('Falla', ''),
                        'responsable': d.get('Responsable', ''),
                        'equipo':      d.get('Equipo', ''),
                        'hora':        d.get('Hora', '00:00'),
                        'tiempo':      (d.get('Tiempo (min)') or d.get('Tiempo (Min)') or d.get('tiempo_minutos') or '0') or '0',
                        'estatus':     ESTATUS_MAP.get(d.get('Estatus', '').lower(), 'rojo'),
                        'comentarios': d.get('Comentarios', ''),
                    }
                    if permitida:
                        filas_permitidas.append(fila_data)
                    else:
                        filas_rechazadas.append(fila_data)

                if not errores:
                    if filas_rechazadas and not es_admin:
                        pendiente_confirmacion = True
                        datos_sesion = json.dumps(filas_permitidas)
                    else:
                        importados = _crear_paros(filas_permitidas)
                        messages.success(request, f"{importados} paro(s) importados correctamente.")
                        return render(request, 'paros_app/importar_paros.html', {
                            'errores':                [],
                            'resumen':                {'importados': importados},
                            'pendiente_confirmacion': False,
                            'datos_sesion':           '',
                            'filas_permitidas':       [],
                            'filas_rechazadas':       [],
                            'columnas':               COLUMNAS,
                            'area_id_param':          area_id_param or '',
                        })

    return render(request, 'paros_app/importar_paros.html', {
        'errores':                  errores,
        'resumen':                  resumen,
        'pendiente_confirmacion':   pendiente_confirmacion,
        'datos_sesion':             datos_sesion,
        'filas_permitidas':         filas_permitidas,
        'filas_rechazadas':         filas_rechazadas,
        'columnas':                 COLUMNAS,
        'area_id_param':            area_id_param or '',
    })